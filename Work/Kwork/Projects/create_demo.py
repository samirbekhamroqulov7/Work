import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

wb = Workbook()

# Стили
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

def auto_width(ws, cols):
    for col in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 35)

# ============================================================
# ЛИСТ 1: СПРАВОЧНИКИ
# ============================================================
ws = wb.active
ws.title = "Справочники"

# Комплекты
ws['A1'] = "СПРАВОЧНИК КОМПЛЕКТОВ"
ws['A1'].font = Font(bold=True, size=14)

headers = ["№", "Код комплекта", "Название комплекта", "Кол-во деталей", "Статус"]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 5)

kits = [
    ["К-001", "Беседка 3x3 стандарт", 8, "Активный"],
    ["К-002", "Беседка 4x4 усиленная", 12, "Активный"],
    ["К-003", "Навес 3x6 односкатный", 6, "Активный"],
    ["К-004", "Навес 4x8 двускатный", 10, "Активный"],
    ["К-005", "Беседка 3x4 с ковкой", 14, "Активный"],
]

for i, kit in enumerate(kits, 4):
    ws.cell(row=i, column=1, value=i-3)
    for j, val in enumerate(kit, 2):
        ws.cell(row=i, column=j, value=val)
style_data(ws, 4, 8, 5)

# Состав комплекта
ws['A11'] = "СОСТАВ КОМПЛЕКТА: К-001 Беседка 3x3 стандарт"
ws['A11'].font = Font(bold=True, size=12)

headers2 = ["№", "Код детали", "Название детали", "Кол-во в комплекте", "Покраска"]
for i, h in enumerate(headers2, 1):
    ws.cell(row=13, column=i, value=h)
style_header(ws, 13, 5)

parts = [
    ["Д-001", "Стойка угловая 80x80x3 L=2200", 4, "Да"],
    ["Д-002", "Балка верхняя 60x40x2 L=3000", 4, "Да"],
    ["Д-003", "Ферма кровельная 40x40x2", 4, "Да"],
    ["Д-004", "Связь диагональная 40x20x2", 8, "Да"],
    ["Д-005", "Пластина опорная 150x150x4", 4, "Нет"],
    ["Д-006", "Обрешётка 20x20x1.5 L=3000", 12, "Да"],
    ["Д-007", "Косынка усиливающая 100x100x3", 8, "Нет"],
    ["Д-008", "Декор. элемент (дуга)", 4, "Да"],
]

for i, part in enumerate(parts, 14):
    ws.cell(row=i, column=1, value=i-13)
    for j, val in enumerate(part, 2):
        ws.cell(row=i, column=j, value=val)
style_data(ws, 14, 21, 5)

auto_width(ws, 5)

# ============================================================
# ЛИСТ 2: СКЛАД МАТЕРИАЛОВ
# ============================================================
ws2 = wb.create_sheet("Склад материалов")

ws2['A1'] = "СКЛАД МАТЕРИАЛОВ (металлопрофиль)"
ws2['A1'].font = Font(bold=True, size=14)

headers3 = ["№", "Профиль", "Размер", "Ед.изм.", "Остаток", "Мин.запас", "Статус"]
for i, h in enumerate(headers3, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 7)

materials = [
    ["Труба квадратная", "80x80x3", "м.п.", 48, 20, ""],
    ["Труба прямоугольная", "60x40x2", "м.п.", 72, 30, ""],
    ["Труба квадратная", "40x40x2", "м.п.", 95, 40, ""],
    ["Труба прямоугольная", "40x20x2", "м.п.", 35, 25, ""],
    ["Труба квадратная", "20x20x1.5", "м.п.", 120, 50, ""],
    ["Лист", "150x150x4", "шт.", 60, 20, ""],
    ["Лист", "100x100x3", "шт.", 45, 20, ""],
    ["Полоса", "40x4", "м.п.", 30, 15, ""],
]

for i, mat in enumerate(materials, 4):
    ws2.cell(row=i, column=1, value=i-3)
    for j, val in enumerate(mat, 2):
        ws2.cell(row=i, column=j, value=val)
    # Статус
    остаток = mat[3]
    мин_запас = mat[4]
    if остаток > мин_запас * 1.5:
        ws2.cell(row=i, column=7, value="✓ Норма")
        ws2.cell(row=i, column=7).fill = green_fill
    elif остаток >= мин_запас:
        ws2.cell(row=i, column=7, value="⚠ Мало")
        ws2.cell(row=i, column=7).fill = yellow_fill
    else:
        ws2.cell(row=i, column=7, value="✗ Дефицит")
        ws2.cell(row=i, column=7).fill = red_fill

style_data(ws2, 4, 11, 7)

# Приход/расход
ws2['A14'] = "ДВИЖЕНИЕ МАТЕРИАЛОВ"
ws2['A14'].font = Font(bold=True, size=12)

headers4 = ["Дата", "Профиль", "Операция", "Кол-во", "Основание"]
for i, h in enumerate(headers4, 1):
    ws2.cell(row=16, column=i, value=h)
style_header(ws2, 16, 5)

moves = [
    ["14.05.2026", "80x80x3", "Приход", 24, "Поставка №45"],
    ["14.05.2026", "60x40x2", "Расход", 12, "Заготовки для К-001"],
    ["13.05.2026", "40x40x2", "Расход", 8, "Заготовки для К-003"],
    ["13.05.2026", "20x20x1.5", "Приход", 50, "Поставка №44"],
    ["12.05.2026", "40x20x2", "Расход", 6, "Заготовки для К-002"],
]

for i, move in enumerate(moves, 17):
    for j, val in enumerate(move, 1):
        ws2.cell(row=i, column=j, value=val)
    if move[2] == "Приход":
        ws2.cell(row=i, column=3).fill = green_fill
    else:
        ws2.cell(row=i, column=3).fill = red_fill

style_data(ws2, 17, 21, 5)
auto_width(ws2, 7)

# ============================================================
# ЛИСТ 3: ЗАГОТОВКИ В РАБОТЕ
# ============================================================
ws3 = wb.create_sheet("Заготовки в работе")

ws3['A1'] = "ЗАГОТОВКИ В РАБОТЕ"
ws3['A1'].font = Font(bold=True, size=14)

headers5 = ["№", "Код", "Название заготовки", "Для детали", "В работе (шт)", "Готово (шт)", "Дата запуска"]
for i, h in enumerate(headers5, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 7)

blanks = [
    ["З-001", "Нарезка трубы 80x80 L=2200", "Д-001 Стойка угловая", 8, 4, "12.05.2026"],
    ["З-002", "Нарезка трубы 60x40 L=3000", "Д-002 Балка верхняя", 12, 8, "12.05.2026"],
    ["З-003", "Сварка фермы 40x40", "Д-003 Ферма кровельная", 4, 2, "13.05.2026"],
    ["З-004", "Нарезка трубы 40x20 L=800", "Д-004 Связь диагональная", 16, 16, "11.05.2026"],
    ["З-005", "Резка листа 150x150x4", "Д-005 Пластина опорная", 8, 8, "11.05.2026"],
    ["З-006", "Нарезка трубы 20x20 L=3000", "Д-006 Обрешётка", 24, 12, "13.05.2026"],
]

for i, blank in enumerate(blanks, 4):
    ws3.cell(row=i, column=1, value=i-3)
    for j, val in enumerate(blank, 2):
        ws3.cell(row=i, column=j, value=val)
style_data(ws3, 4, 9, 7)
auto_width(ws3, 7)

# ============================================================
# ЛИСТ 4: СКЛАД ДЕТАЛЕЙ (БЕЗ ПОКРАСКИ)
# ============================================================
ws4 = wb.create_sheet("Детали без покраски")

ws4['A1'] = "СКЛАД ДЕТАЛЕЙ (БЕЗ ПОКРАСКИ)"
ws4['A1'].font = Font(bold=True, size=14)

headers6 = ["№", "Код детали", "Название", "На складе (шт)", "Дата поступления"]
for i, h in enumerate(headers6, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 5)

unpainted = [
    ["Д-001", "Стойка угловая 80x80x3 L=2200", 4, "13.05.2026"],
    ["Д-002", "Балка верхняя 60x40x2 L=3000", 8, "13.05.2026"],
    ["Д-004", "Связь диагональная 40x20x2", 16, "12.05.2026"],
    ["Д-005", "Пластина опорная 150x150x4", 8, "12.05.2026"],
    ["Д-006", "Обрешётка 20x20x1.5 L=3000", 12, "13.05.2026"],
    ["Д-007", "Косынка усиливающая 100x100x3", 10, "11.05.2026"],
]

for i, part in enumerate(unpainted, 4):
    ws4.cell(row=i, column=1, value=i-3)
    for j, val in enumerate(part, 2):
        ws4.cell(row=i, column=j, value=val)
style_data(ws4, 4, 9, 5)
auto_width(ws4, 5)

# ============================================================
# ЛИСТ 5: СКЛАД ДЕТАЛЕЙ (ПОКРАШЕННЫЕ)
# ============================================================
ws5 = wb.create_sheet("Готовые детали")

ws5['A1'] = "СКЛАД ГОТОВЫХ ДЕТАЛЕЙ (С ПОКРАСКОЙ)"
ws5['A1'].font = Font(bold=True, size=14)

headers7 = ["№", "Код детали", "Название", "На складе (шт)", "Цвет", "Дата покраски"]
for i, h in enumerate(headers7, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 6)

painted = [
    ["Д-001", "Стойка угловая 80x80x3 L=2200", 4, "RAL 8017 (коричн.)", "12.05.2026"],
    ["Д-002", "Балка верхняя 60x40x2 L=3000", 4, "RAL 8017 (коричн.)", "12.05.2026"],
    ["Д-003", "Ферма кровельная 40x40x2", 2, "RAL 9005 (чёрный)", "13.05.2026"],
    ["Д-004", "Связь диагональная 40x20x2", 8, "RAL 8017 (коричн.)", "11.05.2026"],
    ["Д-006", "Обрешётка 20x20x1.5 L=3000", 8, "RAL 9005 (чёрный)", "12.05.2026"],
    ["Д-008", "Декор. элемент (дуга)", 4, "RAL 8017 (коричн.)", "10.05.2026"],
]

for i, part in enumerate(painted, 4):
    ws5.cell(row=i, column=1, value=i-3)
    for j, val in enumerate(part, 2):
        ws5.cell(row=i, column=j, value=val)
style_data(ws5, 4, 9, 6)
auto_width(ws5, 6)

# ============================================================
# ЛИСТ 6: КОМПЛЕКТЫ / ОТГРУЗКА
# ============================================================
ws6 = wb.create_sheet("Комплекты Отгрузка")

ws6['A1'] = "КОМПЛЕКТЫ И ОТГРУЗКА"
ws6['A1'].font = Font(bold=True, size=14)

ws6['A3'] = "ГОТОВНОСТЬ КОМПЛЕКТОВ"
ws6['A3'].font = Font(bold=True, size=12)

headers8 = ["Комплект", "Можно собрать (шт)", "Лимитирующая деталь", "Статус"]
for i, h in enumerate(headers8, 1):
    ws6.cell(row=5, column=i, value=h)
style_header(ws6, 5, 4)

readiness = [
    ["К-001 Беседка 3x3 стандарт", 1, "Д-003 Ферма (нужно 4, есть 2)", "Частично"],
    ["К-002 Беседка 4x4 усиленная", 0, "Д-001 Стойка (нужно 6, есть 4)", "Не готов"],
    ["К-003 Навес 3x6 односкатный", 2, "—", "Готов"],
    ["К-004 Навес 4x8 двускатный", 0, "Д-003 Ферма (нужно 6, есть 2)", "Не готов"],
    ["К-005 Беседка 3x4 с ковкой", 1, "Д-008 Декор (нужно 8, есть 4)", "Частично"],
]

for i, kit in enumerate(readiness, 6):
    for j, val in enumerate(kit, 1):
        ws6.cell(row=i, column=j, value=val)
    status = kit[3]
    if status == "Готов":
        ws6.cell(row=i, column=4).fill = green_fill
    elif status == "Частично":
        ws6.cell(row=i, column=4).fill = yellow_fill
    else:
        ws6.cell(row=i, column=4).fill = red_fill

style_data(ws6, 6, 10, 4)

# Отгрузка
ws6['A13'] = "ЖУРНАЛ ОТГРУЗОК"
ws6['A13'].font = Font(bold=True, size=12)

headers9 = ["Дата", "Комплект", "Кол-во", "Клиент", "Примечание"]
for i, h in enumerate(headers9, 1):
    ws6.cell(row=15, column=i, value=h)
style_header(ws6, 15, 5)

shipments = [
    ["10.05.2026", "К-003 Навес 3x6", 1, "Иванов А.П.", "Доставка"],
    ["08.05.2026", "К-001 Беседка 3x3", 2, "ООО Дачник", "Самовывоз"],
    ["05.05.2026", "К-003 Навес 3x6", 1, "Петров С.В.", "Доставка + монтаж"],
]

for i, ship in enumerate(shipments, 16):
    for j, val in enumerate(ship, 1):
        ws6.cell(row=i, column=j, value=val)
style_data(ws6, 16, 18, 5)
auto_width(ws6, 5)

# ============================================================
# ЛИСТ 7: ПЛАНИРОВАНИЕ
# ============================================================
ws7 = wb.create_sheet("Планирование")

ws7['A1'] = "МОДУЛЬ ПЛАНИРОВАНИЯ"
ws7['A1'].font = Font(bold=True, size=14)

ws7['A3'] = "Введите план производства:"
ws7['A3'].font = Font(bold=True, size=11)

headers10 = ["Комплект", "План (шт)", "Можно сейчас", "Нужно доделать"]
for i, h in enumerate(headers10, 1):
    ws7.cell(row=5, column=i, value=h)
style_header(ws7, 5, 4)

plan = [
    ["К-001 Беседка 3x3 стандарт", 3, 1, 2],
    ["К-002 Беседка 4x4 усиленная", 2, 0, 2],
    ["К-003 Навес 3x6 односкатный", 5, 2, 3],
]

for i, p in enumerate(plan, 6):
    for j, val in enumerate(p, 1):
        ws7.cell(row=i, column=j, value=val)
    ws7.cell(row=i, column=2).fill = subheader_fill  # поле ввода
style_data(ws7, 6, 8, 4)

# Дефицит
ws7['A11'] = "РАСЧЁТ ДЕФИЦИТА ДЕТАЛЕЙ:"
ws7['A11'].font = Font(bold=True, size=12, color="CC0000")

headers11 = ["Деталь", "Нужно всего", "Есть на складе", "Дефицит", "Статус"]
for i, h in enumerate(headers11, 1):
    ws7.cell(row=13, column=i, value=h)
style_header(ws7, 13, 5)

deficit = [
    ["Д-001 Стойка угловая", 20, 4, 16, "✗ Не хватает"],
    ["Д-002 Балка верхняя", 20, 4, 16, "✗ Не хватает"],
    ["Д-003 Ферма кровельная", 22, 2, 20, "✗ Не хватает"],
    ["Д-004 Связь диагональная", 44, 8, 36, "✗ Не хватает"],
    ["Д-005 Пластина опорная", 20, 8, 12, "✗ Не хватает"],
    ["Д-006 Обрешётка", 66, 8, 58, "✗ Не хватает"],
    ["Д-007 Косынка", 20, 10, 10, "✗ Не хватает"],
    ["Д-008 Декор. элемент", 12, 4, 8, "✗ Не хватает"],
]

for i, d in enumerate(deficit, 14):
    for j, val in enumerate(d, 1):
        ws7.cell(row=i, column=j, value=val)
    if d[3] > 0:
        ws7.cell(row=i, column=4).fill = red_fill
        ws7.cell(row=i, column=5).fill = red_fill
    else:
        ws7.cell(row=i, column=4).fill = green_fill
        ws7.cell(row=i, column=5).fill = green_fill

style_data(ws7, 14, 21, 5)

# Дефицит материалов
ws7['A24'] = "РАСЧЁТ ДЕФИЦИТА МАТЕРИАЛОВ:"
ws7['A24'].font = Font(bold=True, size=12, color="CC0000")

headers12 = ["Материал", "Нужно (м.п./шт)", "Есть", "Дефицит", "Статус"]
for i, h in enumerate(headers12, 1):
    ws7.cell(row=26, column=i, value=h)
style_header(ws7, 26, 5)

mat_deficit = [
    ["Труба 80x80x3", "35.2 м.п.", "48 м.п.", "0", "✓ Хватает"],
    ["Труба 60x40x2", "60 м.п.", "72 м.п.", "0", "✓ Хватает"],
    ["Труба 40x40x2", "88 м.п.", "95 м.п.", "0", "✓ Хватает (впритык)"],
    ["Труба 40x20x2", "35.2 м.п.", "35 м.п.", "0.2 м.п.", "✗ Не хватает"],
    ["Труба 20x20x1.5", "198 м.п.", "120 м.п.", "78 м.п.", "✗ Не хватает"],
    ["Лист 150x150x4", "20 шт.", "60 шт.", "0", "✓ Хватает"],
]

for i, m in enumerate(mat_deficit, 27):
    for j, val in enumerate(m, 1):
        ws7.cell(row=i, column=j, value=val)
    if "Не хватает" in m[4]:
        ws7.cell(row=i, column=5).fill = red_fill
    elif "впритык" in m[4]:
        ws7.cell(row=i, column=5).fill = yellow_fill
    else:
        ws7.cell(row=i, column=5).fill = green_fill

style_data(ws7, 27, 32, 5)
auto_width(ws7, 5)

# ============================================================
# ЛИСТ 8: ЖУРНАЛ ОПЕРАЦИЙ
# ============================================================
ws8 = wb.create_sheet("Журнал операций")

ws8['A1'] = "ЖУРНАЛ ОПЕРАЦИЙ"
ws8['A1'].font = Font(bold=True, size=14)

headers13 = ["№", "Дата/Время", "Операция", "Что", "Кол-во", "Откуда", "Куда", "Кто"]
for i, h in enumerate(headers13, 1):
    ws8.cell(row=3, column=i, value=h)
style_header(ws8, 3, 8)

log = [
    [1, "14.05.2026 09:15", "Приход материала", "Труба 80x80x3", "24 м.п.", "Поставщик", "Склад материалов", "Иванов"],
    [2, "14.05.2026 09:30", "Запуск заготовок", "З-001 Нарезка 80x80", "8 шт.", "Склад материалов", "Заготовки в работе", "Петров"],
    [3, "14.05.2026 11:00", "Готовность заготовки", "З-004 Нарезка 40x20", "16 шт.", "Заготовки в работе", "Детали без покраски", "Петров"],
    [4, "14.05.2026 14:00", "Покраска", "Д-001 Стойка угловая", "4 шт.", "Детали без покраски", "Готовые детали", "Сидоров"],
    [5, "14.05.2026 16:00", "Сборка комплекта", "К-003 Навес 3x6", "1 шт.", "Готовые детали", "Комплекты", "Иванов"],
    [6, "14.05.2026 17:00", "Отгрузка", "К-003 Навес 3x6", "1 шт.", "Комплекты", "Клиент: Иванов А.П.", "Иванов"],
]

for i, entry in enumerate(log, 4):
    for j, val in enumerate(entry, 1):
        ws8.cell(row=i, column=j, value=val)
style_data(ws8, 4, 9, 8)
auto_width(ws8, 8)

# Сохранение
filepath = r"c:\Work\Kwork\Projects\Склад_ДЕМО.xlsx"
wb.save(filepath)
print(f"Файл создан: {filepath}")
